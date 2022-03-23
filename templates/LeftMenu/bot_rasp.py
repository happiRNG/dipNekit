import requests
from openpyxl import load_workbook
from io import BytesIO
import datetime
import time


def add_to_db(year, mes, day, group):
    step = 15
    pars = 8
    teacher = {}
    discipline = {}
    auditorium = {}
    id_par = {}
    if day < 10:
        dayy = '0' + str(day)
    else:
        dayy = str(day)
    if mes < 10:
        mess = '0' + str(mes)
    else:
        mess = str(mes)
    req = requests.get("/templates/LeftMenu/rasp/Rasp.-na-" + dayy + "." + mess + "." + str(year) + ".xls")
    row = ''
    if int(req.status_code) == 200:
        op = load_workbook(filename=BytesIO(req.content))
        sheet = op[dayy + "." + mess + "." + str(year)]
        for i in range(1, 550):
            if sheet.cell(i, 2).value == 'Группа:':
                for j in range(3, 17, 4):
                    try:
                        row = sheet.cell(i, j).value
                        if row != group:
                            break
                    except IndexError:
                        pass
                    if row == None:
                        break
                    if i % 2:
                        idd = '1'
                        for ind in range(i + 1, i + step):
                            try:
                                if ind % 2:
                                    teacher[idd] = sheet.cell(ind, j).value
                                else:
                                    idd = sheet.cell(ind, 1).value
                                    discipline[idd] = sheet.cell(ind, j).value
                                    auditorium[idd] = sheet.cell(ind, j + 3).value
                            except IndexError:
                                pass
                        id_par[row] = teacher
                        id_par[row + '_dis'] = discipline
                        id_par[row + '_aud'] = auditorium
                    else:
                        idd = '1'
                        for ind in range(i + 1, i + step):
                            try:
                                if ind % 2:
                                    idd = sheet.cell(ind, 1).value
                                    discipline[idd] = sheet.cell(ind, j).value
                                    auditorium[idd] = sheet.cell(ind, j + 3).value
                                else:
                                    teacher[idd] = sheet.cell(ind, j).value
                            except IndexError:
                                pass
                        id_par[row] = teacher
                        id_par[row + '_dis'] = discipline
                        id_par[row + '_aud'] = auditorium
        return id_par
    return id_par